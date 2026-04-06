import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime


def create_gcp_estimate():
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------
    # 1. 시트 생성 및 이름 설정
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = '고객 입력'
    ws2 = wb.create_sheet('Compute Engine 세부 견적')
    ws3 = wb.create_sheet('Cloud Storage 세부 견적')
    ws4 = wb.create_sheet('Cloud SQL 세부 견적')
    ws5 = wb.create_sheet('Networking 세부 견적')
    ws6 = wb.create_sheet('최종 견적 요약')
    ws7 = wb.create_sheet('단가표')

    # ---------------------------------------------------------
    # 2. 전역 스타일 정의
    # ---------------------------------------------------------
    header_fill   = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True)
    section_fill  = PatternFill(start_color="E8EAED", end_color="E8EAED", fill_type="solid")
    section_font  = Font(bold=True, color="202124")
    input_fill    = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    cud_fill      = PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid")
    total_fill    = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")

    thin = Side(style='thin')
    thin_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    unlocked      = Protection(locked=False)

    fmt_usd = '"$"#,##0.00'
    fmt_krw = '"₩"#,##0'
    fmt_pct = '0%'

    # ---------------------------------------------------------
    # 3. Sheet7: 단가표 (숨김 시트) — asia-northeast3(서울) 기준
    # ---------------------------------------------------------
    # 행 번호 주석: 아래 배치 기준으로 수식에서 참조
    pricing_data = [
        # Row 1
        ["[Compute Engine] — asia-northeast3 기준 On-Demand (USD/시간)"],
        # Row 2
        ["규모",      "스펙",               "Linux단가",  "Windows추가단가"],
        # Row 3
        ["소형",      "e2-standard-2",      0.0860,       0.0920],
        # Row 4
        ["중형",      "e2-standard-4",      0.1720,       0.1840],
        # Row 5
        ["대형",      "e2-standard-8",      0.3440,       0.3680],
        # Row 6
        ["초대형",    "e2-standard-16",     0.6880,       0.7360],
        # Row 7 (빈행)
        [],
        # Row 8
        ["[Cloud SQL] — asia-northeast3 기준 On-Demand (USD/시간)"],
        # Row 9
        ["규모",      "스펙",               "Base단가",   "MSSQL추가단가"],
        # Row 10
        ["소형",      "db-n1-standard-2",   0.1490,       0.2980],
        # Row 11
        ["중형",      "db-n1-standard-4",   0.2980,       0.5960],
        # Row 12
        ["대형",      "db-n1-standard-8",   0.5960,       1.1920],
        # Row 13 (빈행)
        [],
        # Row 14
        ["[기타 서비스] — USD 단가"],
        # Row 15
        ["항목",               "단가",  "",  ""],
        # Row 16
        ["Storage Standard",   0.0230,  "",  ""],
        # Row 17
        ["Storage Snapshot",   0.0260,  "",  ""],
        # Row 18
        ["Network Egress",     0.1200,  "",  ""],
    ]

    for row in pricing_data:
        ws7.append(row)

    # 단가표 헤더 스타일
    for r in [1, 8, 14]:
        ws7.cell(row=r, column=1).font = Font(bold=True, color="1155CC")
    for r in [2, 9, 15]:
        for c in range(1, 5):
            cell = ws7.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = header_font

    ws7.column_dimensions['A'].width = 22
    ws7.column_dimensions['B'].width = 22
    ws7.sheet_state = 'hidden'  # 사용자에게 숨김

    # ---------------------------------------------------------
    # 4. Sheet1: 고객 입력 (비IT 사용자용)
    # ---------------------------------------------------------
    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 28

    # 상단 타이틀
    ws1['A1'] = 'GCP 견적 입력 시트'
    ws1['A1'].font = Font(size=14, bold=True, color="1A73E8")
    ws1['A1'].alignment = left_align

    # (row, 라벨, 기본값, 섹션헤더여부)
    input_rows = [
        (2,  None,                                       None,                       True,  "기본 설정"),
        (3,  "기준 환율 (USD → KRW)",                    1350,                       False, None),
        (4,  "기준 리전",                                 "asia-northeast3",          False, None),
        (5,  "견적 기준일",                               datetime.today().strftime('%Y-%m-%d'), False, None),
        (6,  None,                                       None,                       True,  "서버 및 스토리지"),
        (7,  "서버 규모 (소형/중형/대형/초대형)",          "중형",                     False, None),
        (8,  "서버 대수",                                 2,                          False, None),
        (9,  "사용 OS",                                  "무료 리눅스",              False, None),
        (10, "하루 운영 시간",                             "24h",                     False, None),
        (11, "데이터 저장 용량 (GB)",                     500,                        False, None),
        (12, None,                                       None,                       True,  "데이터베이스 및 네트워크"),
        (13, "DB 규모 (소형/중형/대형)",                  "중형",                     False, None),
        (14, "DB 고가용성(HA) 필요 여부",                  "예",                      False, None),
        (15, "DB 라이선스",                               "무료 PostgreSQL",          False, None),
        (16, "외부 인터넷 전송량 (GB/월)",                 100,                        False, None),
        (17, "백업 필요 여부 (스냅샷)",                    "예",                      False, None),
        (18, "동시 접속자 수 (참고용)",                    1000,                       False, None),
    ]

    for (row, label, default, is_section, section_name) in input_rows:
        if is_section:
            cell_a = ws1.cell(row=row, column=1, value=f"▶  {section_name}")
            cell_a.font   = section_font
            cell_a.fill   = section_fill
            cell_a.alignment = left_align
            ws1.cell(row=row, column=2).fill = section_fill
        else:
            ws1.cell(row=row, column=1, value=label).alignment = left_align
            if default is not None:
                cell_b = ws1.cell(row=row, column=2, value=default)
                cell_b.fill       = input_fill
                cell_b.border     = thin_border
                cell_b.alignment  = center_align
                cell_b.protection = unlocked  # 입력 가능 셀

    # 드롭다운 데이터 유효성 검사
    dv_region  = DataValidation(type="list", formula1='"asia-northeast3,asia-northeast1,asia-east1,us-central1"', allow_blank=False)
    dv_server  = DataValidation(type="list", formula1='"소형,중형,대형,초대형"',     allow_blank=False)
    dv_os      = DataValidation(type="list", formula1='"무료 리눅스,Windows"',       allow_blank=False)
    dv_hours   = DataValidation(type="list", formula1='"8h,12h,24h"',               allow_blank=False)
    dv_db      = DataValidation(type="list", formula1='"소형,중형,대형"',            allow_blank=False)
    dv_yesno   = DataValidation(type="list", formula1='"예,아니오"',                allow_blank=False)
    dv_license = DataValidation(type="list", formula1='"무료 PostgreSQL,MS SQL"',    allow_blank=False)

    ws1.add_data_validation(dv_region);  dv_region.add(ws1["B4"])
    ws1.add_data_validation(dv_server);  dv_server.add(ws1["B7"])
    ws1.add_data_validation(dv_os);      dv_os.add(ws1["B9"])
    ws1.add_data_validation(dv_hours);   dv_hours.add(ws1["B10"])
    ws1.add_data_validation(dv_db);      dv_db.add(ws1["B13"])
    ws1.add_data_validation(dv_yesno);   dv_yesno.add(ws1["B14"])
    ws1.add_data_validation(dv_license); dv_license.add(ws1["B15"])
    ws1.add_data_validation(dv_yesno);   dv_yesno.add(ws1["B17"])

    # Named Range: EXCHANGE_RATE → 고객입력!B3
    dn = DefinedName('EXCHANGE_RATE', attr_text="'고객 입력'!$B$3")
    wb.defined_names.add(dn)

    # 시트 보호 (입력 셀 외 편집 불가)
    ws1.protection.sheet = True
    ws1.freeze_panes = 'A2'

    # ---------------------------------------------------------
    # 5. 세부 견적 시트 공통 헬퍼
    # ---------------------------------------------------------
    def setup_detail_sheet(ws, rows):
        """헤더 1행 + 데이터 행을 추가하고 기본 스타일 적용"""
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35

        header = ["항목", "값 / 수식"]
        ws.append(header)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border

        for r in rows:
            ws.append(r)

        for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
            for cell in row:
                cell.border    = thin_border
                cell.alignment = left_align if cell.column == 1 else center_align

    # ---------------------------------------------------------
    # 6. Sheet2: Compute Engine
    #    행 배치 (헤더=1):
    #    B2=서버규모  B3=스펙  B4=서버대수  B5=OS선택
    #    B6=월운영시간  B7=Base단가  B8=OS단가  B9=월합계(USD)
    # ---------------------------------------------------------
    compute_rows = [
        ["서버 규모",          "='고객 입력'!B7"],
        ["인스턴스 스펙",      "=VLOOKUP(B2, 단가표!$A$3:$D$6, 2, FALSE)"],
        ["서버 대수",          "='고객 입력'!B8"],
        ["OS 선택",            "='고객 입력'!B9"],
        ["월 운영시간 (H)",    '=IF(\'고객 입력\'!B10="8h", 240, IF(\'고객 입력\'!B10="12h", 360, 730))'],
        ["시간당 Linux 단가($)", "=VLOOKUP(B2, 단가표!$A$3:$D$6, 3, FALSE)"],
        ["시간당 OS 추가단가($)", '=IF(B4="Windows", VLOOKUP(B2, 단가표!$A$3:$D$6, 4, FALSE), 0)'],
        # [BUG FIX #1] (B7+B8)×B6×B4 : (Linux단가+OS단가)×월운영시간×서버대수
        ["월 합계 (USD)",      "=(B7+B8)*B6*B4"],
    ]
    setup_detail_sheet(ws2, compute_rows)
    # [BUG FIX #2] 월합계는 B9 (헤더=row1, 데이터시작=row2)
    ws2['B9'].number_format = fmt_usd

    # ---------------------------------------------------------
    # 7. Sheet3: Cloud Storage
    #    B2=저장용량  B3=백업여부  B4=Standard단가  B5=Snapshot단가  B6=월합계(USD)
    # ---------------------------------------------------------
    storage_rows = [
        ["저장 용량 (GB)",         "='고객 입력'!B11"],
        ["백업 여부",              "='고객 입력'!B17"],
        ["Standard 단가 ($/GB)",  '=VLOOKUP("Storage Standard", 단가표!$A$16:$B$18, 2, FALSE)'],
        ["Snapshot 단가 ($/GB)",  '=VLOOKUP("Storage Snapshot", 단가표!$A$16:$B$18, 2, FALSE)'],
        # [BUG FIX #3] B2(용량)×B4(Standard단가) + IF(B3(백업)="예", B2×B5(Snapshot단가), 0)
        ["월 합계 (USD)",          '=B2*B4 + IF(B3="예", B2*B5, 0)'],
    ]
    setup_detail_sheet(ws3, storage_rows)
    # [BUG FIX #4] 월합계는 B6
    ws3['B6'].number_format = fmt_usd

    # ---------------------------------------------------------
    # 8. Sheet4: Cloud SQL
    #    B2=DB규모  B3=스펙  B4=월운영시간  B5=HA여부  B6=라이선스
    #    B7=Base단가  B8=License단가  B9=월합계(USD)
    # ---------------------------------------------------------
    sql_rows = [
        ["DB 규모",               "='고객 입력'!B13"],
        ["인스턴스 스펙",         "=VLOOKUP(B2, 단가표!$A$10:$D$12, 2, FALSE)"],
        ["월 운영시간 (H)",       "=730"],
        ["HA (고가용성) 여부",    "='고객 입력'!B14"],
        ["DB 라이선스",           "='고객 입력'!B15"],
        ["시간당 Base 단가 ($)",  "=VLOOKUP(B2, 단가표!$A$10:$D$12, 3, FALSE)"],
        # [BUG FIX #5] B6=라이선스 텍스트, B5=HA여부 → 라이선스 체크는 B6
        ["시간당 License 단가 ($)", '=IF(B6="MS SQL", VLOOKUP(B2, 단가표!$A$10:$D$12, 4, FALSE), 0)'],
        # [BUG FIX #6] (B7+B8)×B4×HA계수 : (Base단가+License단가)×월운영시간×HA배수
        ["월 합계 (USD)",         '=(B7+B8)*B4*IF(B5="예", 2, 1)'],
    ]
    setup_detail_sheet(ws4, sql_rows)
    # [BUG FIX #7] 월합계는 B9
    ws4['B9'].number_format = fmt_usd

    # ---------------------------------------------------------
    # 9. Sheet5: Networking
    #    B2=전송량  B3=GB당단가  B4=월합계(USD)
    # ---------------------------------------------------------
    net_rows = [
        ["외부 전송량 (GB/월)",   "='고객 입력'!B16"],
        ["GB당 Egress 단가 ($)", '=VLOOKUP("Network Egress", 단가표!$A$16:$B$18, 2, FALSE)'],
        # [BUG FIX #8] B2×B3 (B1은 헤더행이므로 참조 불가)
        ["월 합계 (USD)",         "=B2*B3"],
    ]
    setup_detail_sheet(ws5, net_rows)
    # [BUG FIX #9] 월합계는 B4
    ws5['B4'].number_format = fmt_usd

    # ---------------------------------------------------------
    # 10. Sheet6: 최종 견적 요약 (IT 전문가용 제출 양식)
    # ---------------------------------------------------------
    ws6.column_dimensions['A'].width = 28
    ws6.column_dimensions['B'].width = 22
    ws6.column_dimensions['C'].width = 10
    ws6.column_dimensions['D'].width = 16
    ws6.column_dimensions['E'].width = 20
    ws6.column_dimensions['F'].width = 20

    # Row 1: 문서 제목
    ws6.merge_cells('A1:F1')
    ws6['A1'] = 'Google Cloud Platform 공식 견적서'
    ws6['A1'].font      = Font(size=16, bold=True, color="1A73E8")
    ws6['A1'].alignment = center_align
    ws6['A1'].fill      = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    # Row 2: 메타 정보
    ws6['A2'] = '견적 기준일'
    ws6['B2'] = "='고객 입력'!B5"
    ws6['D2'] = '적용 환율 (KRW/USD)'
    ws6['E2'] = '=EXCHANGE_RATE'
    ws6['F2'] = '원'
    for cell in [ws6['A2'], ws6['D2']]:
        cell.font = Font(bold=True)

    # Row 3: 리전
    ws6['A3'] = '선택 리전'
    ws6['B3'] = "='고객 입력'!B4"
    ws6['A3'].font = Font(bold=True)

    # Row 4: 빈 행
    ws6.append([])

    # Row 5: 테이블 헤더
    col_headers = ["서비스명 (GCP)", "스펙 / 상세", "수량", "시간당 단가 (USD)", "월 비용 (USD)", "월 비용 (KRW)"]
    ws6.append(col_headers)
    for cell in ws6[5]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border

    # Row 6: Compute Engine
    # [BUG FIX #10] 월합계 참조를 !B9로 수정 (기존 !B8은 OS단가 셀)
    ws6.append([
        "Compute Engine",
        "='Compute Engine 세부 견적'!B3",   # 인스턴스 스펙
        "='Compute Engine 세부 견적'!B4",   # 서버 대수
        "='Compute Engine 세부 견적'!B7",   # 시간당 Linux 단가
        "='Compute Engine 세부 견적'!B9",   # 월합계(USD) ← BUG FIX
        "=E6*EXCHANGE_RATE",
    ])

    # Row 7: Cloud Storage
    ws6.append([
        "Cloud Storage",
        "Standard + Snapshot",
        "1 식",
        "-",
        "='Cloud Storage 세부 견적'!B6",    # 월합계(USD)
        "=E7*EXCHANGE_RATE",
    ])

    # Row 8: Cloud SQL
    # [BUG FIX #11] 월합계 참조를 !B9로 수정 (기존 !B8은 License단가 셀)
    ws6.append([
        "Cloud SQL",
        "='Cloud SQL 세부 견적'!B3",        # 인스턴스 스펙
        "1 식",
        "='Cloud SQL 세부 견적'!B7",        # 시간당 Base 단가
        "='Cloud SQL 세부 견적'!B9",        # 월합계(USD) ← BUG FIX
        "=E8*EXCHANGE_RATE",
    ])

    # Row 9: Networking
    ws6.append([
        "Network Egress",
        "인터넷 외부 전송",
        "='Networking 세부 견적'!B2",       # 전송량(GB)
        "='Networking 세부 견적'!B3",       # GB당 단가
        "='Networking 세부 견적'!B4",       # 월합계(USD)
        "=E9*EXCHANGE_RATE",
    ])

    # Row 10: 종량제 소계
    ws6.append([
        "종량제 (On-Demand) 소계", "", "", "",
        "=SUM(E6:E9)",
        "=SUM(F6:F9)",
    ])
    for c in range(1, 7):
        cell = ws6.cell(row=10, column=c)
        cell.font   = Font(bold=True)
        cell.fill   = PatternFill(start_color="D2E3FC", end_color="D2E3FC", fill_type="solid")
        cell.border = thin_border

    # Row 11: 빈행
    ws6.append([])

    # Row 12: CUD 섹션 타이틀
    ws6.merge_cells('A12:F12')
    ws6['A12'] = '장기 약정 할인 (Committed Use Discount, CUD)'
    ws6['A12'].font      = Font(bold=True, color="B45309")
    ws6['A12'].fill      = cud_fill
    ws6['A12'].alignment = center_align

    # Row 13: CUD 컬럼 헤더
    cud_headers = ["약정 기간", "예상 할인율 (수정 가능)", "", "", "할인 적용 월비용 (USD)", "할인 적용 월비용 (KRW)"]
    ws6.append(cud_headers)
    for cell in ws6[13]:
        cell.fill      = cud_fill
        cell.font      = Font(bold=True)
        cell.border    = thin_border
        cell.alignment = center_align

    # Row 14: 1년 약정 (할인율 셀 직접 수정 가능)
    ws6.append(["1년 약정 (1-Year CUD)", 0.28, "", "", "=E10*(1-B14)", "=F10*(1-B14)"])
    ws6['B14'].number_format = fmt_pct
    ws6['B14'].fill          = input_fill   # 수정 가능 강조
    for cell in ws6[14]:
        cell.border    = thin_border
        cell.alignment = center_align

    # Row 15: 3년 약정
    ws6.append(["3년 약정 (3-Year CUD)", 0.46, "", "", "=E10*(1-B15)", "=F10*(1-B15)"])
    ws6['B15'].number_format = fmt_pct
    ws6['B15'].fill          = input_fill
    for cell in ws6[15]:
        cell.border    = thin_border
        cell.alignment = center_align

    # Row 16: 빈행
    ws6.append([])

    # Row 17: 연간 합계 (3년 약정 기준)
    ws6.merge_cells('A17:D17')
    ws6['A17'] = '총 연간 예상 비용 (3년 약정 기준)'
    ws6['A17'].font      = Font(bold=True, size=12)
    ws6['A17'].fill      = total_fill
    ws6['A17'].alignment = center_align
    ws6['E17'] = '=E15*12'
    ws6['F17'] = '=F15*12'
    ws6['E17'].font = Font(bold=True, size=12, color="137333")
    ws6['F17'].font = Font(bold=True, size=12, color="137333")
    ws6['E17'].fill = total_fill
    ws6['F17'].fill = total_fill
    for c in range(1, 7):
        ws6.cell(row=17, column=c).border = thin_border

    # Row 18: 빈행
    ws6.append([])

    # Row 19: 면책 문구
    ws6.merge_cells('A19:F19')
    ws6['A19'] = '※ 본 견적은 GCP 공식 가격 기준이며, 실제 청구 금액은 사용 패턴에 따라 달라질 수 있습니다.'
    ws6['A19'].font      = Font(italic=True, color="5F6368", size=9)
    ws6['A19'].alignment = left_align

    # 숫자 포맷 일괄 적용 (rows 6~10, 14~15, 17)
    for row_idx in range(6, 11):
        ws6.cell(row=row_idx, column=4).number_format = fmt_usd
        ws6.cell(row=row_idx, column=5).number_format = fmt_usd
        ws6.cell(row=row_idx, column=6).number_format = fmt_krw

    for row_idx in [14, 15]:
        ws6.cell(row=row_idx, column=5).number_format = fmt_usd
        ws6.cell(row=row_idx, column=6).number_format = fmt_krw

    ws6['E17'].number_format = fmt_usd
    ws6['F17'].number_format = fmt_krw

    # 조건부 서식: 월비용(USD) $500 초과 시 빨간색 강조
    red_font = Font(color="D93025", bold=True)
    rule = CellIsRule(operator='greaterThan', formula=['500'], font=red_font)
    ws6.conditional_formatting.add('E6:E9', rule)

    # 헤더 고정 및 시트 보호 해제 (견적서는 편집 허용)
    ws6.freeze_panes = 'A6'

    # ---------------------------------------------------------
    # 11. 파일 저장
    # ---------------------------------------------------------
    filename = "gcp_estimate.xlsx"
    wb.save(filename)
    print(f"✅ '{filename}' 파일이 성공적으로 생성되었습니다.")


if __name__ == "__main__":
    create_gcp_estimate()
